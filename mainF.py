import tkinter as tk
from PIL import Image, ImageTk
import cv2
from deepface import DeepFace
import subprocess
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from pathlib import Path
import pyttsx3
import threading
import face_recognition
from deepface import DeepFace
from collections import Counter
from PIL import Image, ImageTk
import os






class FaceRecUI:
    def __init__(self):
        os.makedirs("db", exist_ok=True)
        self.main_window = tk.Tk()
        self.main_window.title("FaceRec")
        self.main_window.geometry("1250x520+300+100")
        self.main_window.configure(bg="#1f1f1f")

        self._init_camera_ui()
        self._init_buttons()
        self._load_logo()
        self.log_dir = Path("logs")
        self.log_dir.mkdir(exist_ok=True)
        self.excel_path = self.log_dir / f"attendance_{datetime.now().strftime('%Y%m%d')}.xlsx"

        self._init_excel()

        self.tts_lock = threading.Lock()
        self._init_tts()

        self.cap = cv2.VideoCapture(0)
        if not self.cap.isOpened():
            print("❌ دوربین در دسترس نیست.")
        else:
            self.update_camera()

    def _init_camera_ui(self):
        """لیبل نمایش تصویر وبکم"""
        self.cam_label = tk.Label(self.main_window, bg="#1f1f1f")
        self.cam_label.place(x=10, y=10, width=700, height=500)

    def update_camera(self):
        ret, frame = self.cap.read()
        if not ret:
            print("⚠️ Frame capture failed.")
            return

        # resize برای پردازش سریع‌تر
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        # شناسایی چهره‌ها
        face_locations = face_recognition.face_locations(rgb_small_frame)

        # تشخیص احساس فقط روی یک فریم در هر چندتا انجام بشه
        if not hasattr(self, 'frame_counter'):
            self.frame_counter = 0

        if self.frame_counter % 5 == 0 and face_locations:
            age, emotion = self.detect_age_emotion(rgb_small_frame)
            self.last_age = age
            self.last_emotion = emotion
        else:
            age = getattr(self, 'last_age', "Unknown")
            emotion = getattr(self, 'last_emotion', "neutral")

        self.frame_counter += 1

        # بازگشت مقیاس مختصات چهره به اندازه اصلی
        face_locations = [(top * 4, right * 4, bottom * 4, left * 4) for (top, right, bottom, left) in face_locations]

        # رسم کادر و اطلاعات
        for (top, right, bottom, left) in face_locations:
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        cv2.putText(frame, f"Age: {age}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)
        cv2.putText(frame, f"Emotion: {emotion}", (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 0, 0), 2)

        # نمایش در رابط گرافیکی
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(rgb_frame)
        imgtk = ImageTk.PhotoImage(image=img)

        self.cam_label.imgtk = imgtk
        self.cam_label.configure(image=imgtk)

        self.cam_label.after(20, self.update_camera)

    def _init_buttons(self):
        """ساخت دکمه‌های سمت راست"""
        # دکمه Login
        self.login_btn = tk.Button(self.main_window, text="Login", font=("Arial", 16),
                                   bg="green", fg="white", width=20, command=self.dummy_login)
        self.login_btn.place(x=800, y=350)
        self.login_btn.bind("<Enter>", lambda e: self._on_hover(e, "darkgreen"))
        self.login_btn.bind("<Leave>", lambda e: self._on_hover(e, "green"))

        # دکمه Register
        self.register_btn = tk.Button(self.main_window, text="Register New User", font=("Arial", 16),
                                      bg="white", fg="black", width=20, command=self.dummy_register)
        self.register_btn.place(x=800, y=400)
        self.register_btn.bind("<Enter>", lambda e: self._on_hover(e, "lightgray"))
        self.register_btn.bind("<Leave>", lambda e: self._on_hover(e, "white"))

        # متن بالای دکمه‌ها
        self.title_label = tk.Label(self.main_window, text="",
                                    font=("Arial", 22, "bold"), fg="white", bg="#1f1f1f")
        self.title_label.place(x=800, y=100)

    def _on_hover(self, event, color):
        event.widget.config(bg=color)

    def detect_age_emotion(self, frame):
        try:
            results = DeepFace.analyze(frame, actions=['age', 'emotion'], enforce_detection=False, silent=True)
            age = results[0].get('age', 'Unknown')
            emotion = results[0].get('dominant_emotion', 'neutral').lower()
            return age, emotion
        except Exception as e:
            print("Emotion/Age detection error:", e)
            return "Unknown", "neutral"

    def _init_tts(self):
        try:
            self.tts_engine = pyttsx3.init()
            self.tts_engine.setProperty('rate', 150)
            self.tts_engine.setProperty('volume', 1.0)

            # انتخاب صدای زن اگر موجود باشه
            for voice in self.tts_engine.getProperty('voices'):
                if "female" in voice.name.lower():
                    self.tts_engine.setProperty('voice', voice.id)
                    break
        except Exception as e:
            print("❌ خطا در راه‌اندازی TTS:", e)


    def speak(self, text):
        def _speak():
            with self.tts_lock:
                try:
                    self.tts_engine.say(text)
                    self.tts_engine.runAndWait()
                except Exception as e:
                    print("🔈 خطا در TTS:", e)

        threading.Thread(target=_speak, daemon=True).start()


    def dummy_login(self):
        # ذخیره فریم فعلی
        ret, frame = self.cap.read()
        if not ret:
            print("❌ Frame capture failed.")
            return

        # ذخیره تصویر موقت
        os.makedirs("temp", exist_ok=True)
        temp_img_path = "temp/temp.jpg"
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        cv2.imwrite(temp_img_path, rgb_frame)

        # اجرای face_recognition CLI
        try:
            cmd = ['face_recognition', '--tolerance', '0.6', 'db/', temp_img_path]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)

            if result.returncode != 0 or not result.stdout.strip():
                print("❌ Face not recognized.")
                self._show_popup("Login Failed", "Face not recognized.")
                return

            parts = result.stdout.strip().split(',')
            if len(parts) < 2:
                print("⚠️ Unexpected result:", result.stdout)
                self._show_popup("Login Failed", "Invalid face recognition result.")
                return

            name = parts[1].strip()

            # تشخیص احساس
            age, emotion = self.detect_age_emotion(rgb_frame)

            # پیام خوش‌آمد
            message = {
                'happy': f"Welcome {name}! You look happy today!",
                'sad': f"Welcome {name}. We hope your day gets better.",
                'angry': f"Welcome {name}. We appreciate your patience.",
            }.get(emotion, f"Welcome {name}!")

            self._show_popup("Login Success", message)

            print(f"[{datetime.now()}] {name} logged in with emotion: {emotion}")
            self._log_attendance(name, emotion)
            self.speak(message)



        except Exception as e:
            print("❌ Login error:", str(e))
            self._show_popup("Login Failed", str(e))

    def dummy_register(self):
        # پنجره جدید برای ثبت‌نام
        self.register_window = tk.Toplevel(self.main_window)
        self.register_window.title("Register New User")
        self.register_window.geometry("1100x550+350+150")
        self.register_window.configure(bg="#1f1f1f")

        # تصویر دوربین
        self.capture_label = tk.Label(self.register_window, bg="#1f1f1f")
        self.capture_label.place(x=20, y=20, width=700, height=500)

        # گرفتن یک فریم برای نمایش ثابت
        ret, frame = self.cap.read()
        if ret:
            self.registration_frame = frame.copy()
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            rgb = cv2.resize(rgb, (700, 500))
            img = Image.fromarray(rgb)
            imgtk = ImageTk.PhotoImage(image=img)
            self.capture_label.imgtk = imgtk
            self.capture_label.configure(image=imgtk)

        # فیلد ورودی نام
        tk.Label(self.register_window, text="Enter Name:", font=("Poppins", 14), fg="white", bg="#1f1f1f").place(x=750,
                                                                                                               y=100)
        self.name_entry = tk.Text(self.register_window, height=1, width=25, font=("Arial", 14))
        self.name_entry.place(x=750, y=140)

        # دکمه ثبت
        self.save_btn = tk.Button(self.register_window, text="Save", font=("Arial", 14), bg="green", fg="white",
                                  command=self._save_new_user)
        self.save_btn.place(x=750, y=200)

        self.save_btn.bind("<Enter>", lambda e: self._on_hover(e, "darkgreen"))
        self.save_btn.bind("<Leave>", lambda e: self._on_hover(e, "green"))

        # دکمه انصراف
        self.tryAgain_btn = tk.Button(self.register_window, text="Try again", font=("Arial", 14), bg="red", fg="white",
                  command=self.register_window.destroy)

        self.tryAgain_btn.place(x=830, y=200)

        self.tryAgain_btn.bind("<Enter>", lambda e: self._on_hover(e, "darkred"))
        self.tryAgain_btn.bind("<Leave>", lambda e: self._on_hover(e, "red"))

    def _save_new_user(self):
        name = self.name_entry.get("1.0", "end-1c").strip().lower().replace(" ", "_")
        if not name:
            self._show_popup("Error", "Please enter a valid name.")
            return

        save_path = os.path.join("db", f"{name}.jpg")

        try:
            cv2.imwrite(save_path, self.registration_frame)
            self._show_popup("Success", f"{name} registered successfully.")
            self.register_window.destroy()
        except Exception as e:
            print(f"Error saving user: {e}")
            self._show_popup("Error", "Failed to save image.")

    def _show_popup(self, title, message):
        popup = tk.Toplevel(self.main_window)
        popup.title(title)
        popup.geometry("400x200")
        popup.configure(bg="white")

        tk.Label(popup, text=message, font=("Arial", 14), wraplength=350, bg="white").pack(pady=40)
        tk.Button(popup, text="OK", command=popup.destroy, bg="green", fg="white").pack()

    def _init_excel(self):
        """اگر فایل موجود نیست، ایجادش کن"""
        if not self.excel_path.exists():
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Name", "Emotion", "Time In", "Time Out", "Duration"])
            wb.save(self.excel_path)

    def _log_attendance(self, name, emotion):
        now = datetime.now()
        time_str = now.strftime("%I:%M:%S %p").lower()

        # بارگذاری فایل اکسل
        wb = load_workbook(self.excel_path)
        sheet = wb.active

        updated = False

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == name and row[3].value is None:
                try:
                    time_in_str = row[2].value
                    time_in = datetime.strptime(time_in_str, "%I:%M:%S %p").time()
                    time_in_dt = datetime.combine(now.date(), time_in)

                    # اگر logout بعد از نیمه‌شب باشه
                    if now.time() < time_in:
                        time_in_dt -= timedelta(days=1)

                    duration = now - time_in_dt
                    dur_str = str(duration).split(".")[0]  # فقط HH:MM:SS

                    row[3].value = time_str  # Time Out
                    row[4].value = dur_str  # Duration
                    updated = True
                    break
                except Exception as e:
                    print(f"Duration calc error: {e}")

        if not updated:
            # ثبت ورود جدید
            sheet.append([name, emotion, time_str, None, None])

        wb.save(self.excel_path)

    def _load_logo(self):
        try:
            logo_path = os.path.join("logo", "image1.png")  # یا "logo.png"
            if not os.path.exists(logo_path):
                print("❌ فایل لوگو پیدا نشد.")
                return

            logo_img = Image.open(logo_path)
            logo_img = logo_img.resize((410, 245))  # سایز دلخواه
            self.logo_photo = ImageTk.PhotoImage(logo_img)

            self.logo_label = tk.Label(self.main_window, image=self.logo_photo, bg="#1f1f1f")
            self.logo_label.place(x=730, y=75)  # موقعیت دلخواه بالا سمت راست
        except Exception as e:
            print(f"خطا در بارگذاری لوگو: {e}")

    def start(self):
        self.main_window.mainloop()

    def cleanup(self):
        if hasattr(self, 'cap') and self.cap.isOpened():
            self.cap.release()

if __name__ == "__main__":
    app = FaceRecUI()
    try:
        app.start()
    finally:
        app.cleanup()

