import os
import cv2
import tkinter as tk
import util
import threading
import time
import logging
import subprocess
import face_recognition
from datetime import datetime, timedelta
from PIL import Image, ImageTk
from tkinter import Canvas
import pyttsx3
from deepface import DeepFace
from openpyxl import Workbook, load_workbook
from pathlib import Path

# Configure logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("app_debug.log"),
        logging.StreamHandler()
    ]
)


class App:
    def __init__(self):
        # Initialize main window
        self.main_window = tk.Tk()
        self.main_window.withdraw()

        # Initialize TTS engine early
        self.engine_lock = threading.Lock()
        self._init_tts_engine()

        # Setup splash screen
        self._setup_splash_screen()
        self._update_splash("Starting application...", force=True)

        # Initialize components
        self._init_window_properties()
        self._update_splash("Loading interface...", delay=0.5)
        self._init_canvas()
        self._update_splash("Initializing camera...", delay=0.5)
        self._init_camera()
        self._update_splash("Creating buttons...", delay=0.5)
        self._init_buttons()
        self._update_splash("Preparing database...", delay=0.5)
        self._init_database()
        self._update_splash("Setting up logging...", delay=0.5)
        self._init_logging_system()
        self._update_splash("Completed!", delay=0.5)

        # Final setup
        self.splash.destroy()
        self.main_window.deiconify()

    def _init_tts_engine(self):
        """Initialize TTS engine with thread safety"""
        with self.engine_lock:
            if not hasattr(self, 'engine'):
                self.engine = pyttsx3.init()
                voices = self.engine.getProperty('voices')
                for voice in voices:
                    if "female" in voice.name.lower():
                        self.engine.setProperty('voice', voice.id)
                        break
                self.engine.setProperty('rate', 150)
                self.engine.setProperty('volume', 1.0)
                self.engine.startLoop(False)

    def speak(self, text):
        """Thread-safe text-to-speech"""

        def _speak():
            with self.engine_lock:
                try:
                    self.engine.say(text)
                    self.engine.iterate()
                except RuntimeError:
                    self.engine.endLoop()
                    self.engine.startLoop(False)
                    self.engine.say(text)
                    self.engine.iterate()
                except Exception as e:
                    logging.error(f"TTS error: {str(e)}")

        threading.Thread(target=_speak, daemon=True).start()

    def _setup_splash_screen(self):
        """Create splash screen with guaranteed visibility"""
        self.splash = tk.Toplevel()
        self.splash.geometry("1200x750+500+300")
        self.splash.overrideredirect(True)
        self.splash.configure(bg="#000000")
        self.splash.attributes("-topmost", True)

        # Load splash image
        splash_img = Image.open("splashImage/splashimg.jpg").resize((1000, 600))
        self.splash_photo = ImageTk.PhotoImage(splash_img)
        tk.Label(self.splash, image=self.splash_photo, bg="#000000").pack(pady=10)

        # Status label
        self.status_var = tk.StringVar()
        self.status_var.set("Version 1.2.0 | Starting...")
        tk.Label(
            self.splash,
            textvariable=self.status_var,
            font=("Arial", 14, "bold"),
            bg="#1f1f1f",
            fg="white",
            padx=20,
            pady=8
        ).pack(side="bottom", pady=22)

        self.splash.update_idletasks()
        self.splash.update()

    def _update_splash(self, message, delay=0.5, force=False):
        """Update splash message with guaranteed visibility"""
        if hasattr(self, 'splash'):
            self.status_var.set(f"Version 1.2.0 | {message}")
            self.splash.update_idletasks()
            self.splash.update()
            time.sleep(delay)

    def _init_window_properties(self):
        """Initialize main window properties"""
        self.main_window.geometry("1250x520+350+100")
        self.main_window.configure(bg="#1f1f1f")
        self.main_window.title("FaceRec")

    def _init_buttons(self):
        """Initialize all interface buttons"""
        # Login button
        self.login_button_main_window = util.get_button(
            self.main_window,
            'Login',
            'green',
            self.login
        )
        self.login_button_main_window.place(x=750, y=300)

        # Register button
        self.register_new_user_button_main_window = util.get_button(
            self.main_window,
            'Register new user',
            'white',
            self.register_new_user,
            fg='black'
        )
        self.register_new_user_button_main_window.place(x=750, y=400)

        # Hover effects
        self.login_button_main_window.bind("<Enter>", lambda e: self.on_enter(e, "darkgreen"))
        self.login_button_main_window.bind("<Leave>", lambda e: self.on_leave(e, "green"))
        self.register_new_user_button_main_window.bind("<Enter>", lambda e: self.on_enter(e, "lightgray"))
        self.register_new_user_button_main_window.bind("<Leave>", lambda e: self.on_leave(e, "white"))

        self.main_window.update()

    def _init_canvas(self):
        """Initialize canvas components including logo"""
        self._update_splash("Setting up interface...", delay=2)

        self.canvas = Canvas(self.main_window, width=1100, height=600, bd=0, highlightthickness=0)
        self.canvas.grid(row=0, column=0, columnspan=2, pady=20)
        self.canvas.configure(bg="#1f1f1f")

        # Load and display logo (top-right)
        image_path = os.path.join(os.path.dirname(__file__), "logo", "image1.png")
        if os.path.exists(image_path):
            try:
                self._update_splash("Loading logo...", delay=1.6)
                self.img_file = Image.open(image_path)
                self.img_file = self.img_file.resize((370, 220))
                self.img = ImageTk.PhotoImage(self.img_file)

                # Calculate position (top-right)
                window_width = 1100
                image_width = 370
                x_position = window_width - image_width
                self.canvas.create_image(x_position, 0, anchor=tk.NW, image=self.img)

                self.main_window.update()
            except Exception as e:
                self._update_splash(f"Logo error: {str(e)}", delay=1)
                logging.error(f"Error loading logo: {e}")
        else:
            self._update_splash("Logo not found!", delay=0.5)
            logging.error(f"Logo file not found at: {image_path}")

    def _init_camera(self):
        """Initialize camera with visual feedback"""
        self.webcam_label = util.get_img_label(self.main_window)
        self.webcam_label.place(x=10, y=10, width=700, height=500)
        self.webcam_label.configure(bg="#1f1f1f")
        self.add_webcam(self.webcam_label)

    def _init_database(self):
        """Initialize database directory"""
        self.db_dir = './db'
        if not os.path.exists(self.db_dir):
            os.mkdir(self.db_dir)

    def _init_logging_system(self):
        """Initialize the logging system with proper error handling"""
        self.log_dir = Path("./attendance_logs")
        self.log_dir.mkdir(exist_ok=True)
        self.excel_path = self.log_dir / f"attendance_{datetime.now().strftime('%Y%m%d')}.xlsx"

        try:
            if self.excel_path.exists():
                try:
                    # Open with read-only first to check integrity
                    with open(self.excel_path, 'rb') as f:
                        load_workbook(f)
                    self.workbook = load_workbook(self.excel_path)
                except Exception as e:
                    # Create backup of corrupt file
                    corrupt_time = datetime.now().strftime("%Y%m%d_%H%M%S")
                    corrupt_path = self.log_dir / f"corrupt_{corrupt_time}.xlsx"
                    self.excel_path.rename(corrupt_path)
                    logging.warning(f"Moved corrupt file to: {corrupt_path}")
                    self._create_new_workbook()
            else:
                self._create_new_workbook()
        except Exception as e:
            logging.error(f"Excel initialization error: {e}")
            self._create_emergency_log()

    def _create_new_workbook(self):
        """Create fresh workbook with proper headers"""
        try:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Name", "Emotion", "Time In", "Time Out", "Duration"])
            self._safe_excel_save()
            logging.info(f"Created new Excel file at: {self.excel_path}")
        except Exception as e:
            logging.error(f"Failed to create new workbook: {e}")
            raise

    def _create_emergency_log(self):
        """Fallback when primary logging fails"""
        self.excel_path = self.log_dir / "emergency_attendance.xlsx"
        try:
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(["Name", "Emotion", "Time In", "Time Out", "Duration"])
            self._safe_excel_save()
        except Exception as e:
            logging.error(f"Critical logging failure: {e}")


    def _safe_excel_save(self):
        """Atomic Excel file save with retry logic"""
        for attempt in range(3):
            try:
                temp_path = self.excel_path.with_suffix('.tmp')

                # Save to temporary file
                self.workbook.save(temp_path)

                # Replace original file
                if self.excel_path.exists():
                    self.excel_path.unlink()
                temp_path.rename(self.excel_path)
                return True
            except Exception as e:
                logging.error(f"Save attempt {attempt + 1} failed: {e}")
                time.sleep(0.5)
        return False


    def _log_attendance(self, name, emotion):
        """Thread-safe attendance logging with enhanced error handling"""
        for attempt in range(3):
            try:
                now = datetime.now()
                time_str = now.strftime("%I:%M:%S %p").lower()  # Format: "11:23:20 pm"

                # 1. Ensure directories exist
                self.log_dir.mkdir(parents=True, exist_ok=True)

                # 2. Initialize workbook if needed
                if not self.excel_path.exists():
                    self._create_new_workbook()

                # 3. Load workbook
                try:
                    self.workbook = load_workbook(self.excel_path)
                except Exception as e:
                    logging.error(f"Error loading workbook: {e}")
                    self._handle_corrupt_excel()
                    self.workbook = load_workbook(self.excel_path)

                self.sheet = self.workbook.active

                # 4. Find existing record
                record_updated = False
                for idx, row in enumerate(self.sheet.iter_rows(min_row=2), start=2):
                    if row[0].value == name and row[3].value is None:
                        try:
                            # Parse Time In (already in 12-hour format)
                            time_in_str = row[2].value  # e.g., "09:30:45 am"
                            time_in = datetime.strptime(time_in_str, "%I:%M:%S %p").time()
                            time_in_dt = datetime.combine(now.date(), time_in)

                            # Handle overnight case (e.g., Time Out after midnight)
                            if now.time() < time_in:
                                time_in_dt -= timedelta(days=1)

                            # Calculate duration
                            duration = now - time_in_dt
                            hours, rem = divmod(int(duration.total_seconds()), 3600)
                            minutes, seconds = divmod(rem, 60)
                            duration_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"  # Duration in HH:MM:SS

                            # Update record (Time Out and Duration)
                            self.sheet.cell(row=idx, column=4, value=time_str)  # Time Out
                            self.sheet.cell(row=idx, column=5, value=duration_str)  # Duration
                            record_updated = True
                            break
                        except Exception as e:
                            logging.warning(f"Time calculation error: {e}")
                            continue

                # 5. If no existing record, add new entry
                if not record_updated:
                    self.sheet.append([name, emotion, time_str, None, None])  # Time In only

                # 6. Save changes
                if not self._safe_excel_save():
                    raise Exception("Failed to save workbook after 3 attempts")

                logging.info(f"Logged attendance for {name}")
                return True

            except Exception as e:
                logging.error(f"Attempt {attempt + 1} failed: {e}")
                if attempt == 2:  # Final attempt
                    self._create_emergency_log()
                    logging.critical(f"Logged emergency entry for {name}")
                time.sleep(0.5)
        return False


    def _log_to_text_backup(self, name, emotion, time_str):
        """Fallback logging when Excel fails"""
        backup_path = Path("./attendance_logs/emergency_log.txt")
        try:
            backup_path.parent.mkdir(exist_ok=True, parents=True)
            with open(backup_path, "a") as f:
                f.write(f"{datetime.now().isoformat()}|{name}|{emotion}|{time_str}\n")
        except Exception as e:
            logging.error(f"Backup logging failed: {str(e)}")

    def on_enter(self, event, color):
        event.widget.config(bg=color)

    def on_leave(self, event, color):
        event.widget.config(bg=color)

    def add_webcam(self, label):
        if 'cap' not in self.__dict__:
            self.cap = cv2.VideoCapture(0)

        self._label = label
        self.process_webcam()

    def detect_age_emotion(self, frame):
        try:
            # Analyze frame with DeepFace
            result = DeepFace.analyze(
                frame,
                actions=['age', 'emotion'],
                enforce_detection=False,
                silent=True
            )

            # Safely extract age and emotion
            age = result[0].get('age', 'Unknown')
            emotion = result[0].get('dominant_emotion', 'neutral').lower()

            # Update emotion buffer
            if not hasattr(self, 'emotion_buffer'):
                self.emotion_buffer = []
                self.buffer_size = 5

            self.emotion_buffer.append(emotion)
            if len(self.emotion_buffer) > self.buffer_size:
                self.emotion_buffer.pop(0)

            # Get most frequent emotion
            from collections import Counter
            if self.emotion_buffer:
                final_emotion = Counter(self.emotion_buffer).most_common(1)[0][0]
            else:
                final_emotion = 'neutral'

            return age, final_emotion

        except Exception as e:
            logging.error(f"Error in emotion detection: {str(e)}")
            return "Unknown", "neutral"

    def process_webcam(self):
        ret, frame = self.cap.read()

        if ret:
            if not hasattr(self, 'frame_counter'):
                self.frame_counter = 0

            # Initialize variables
            if not hasattr(self, 'last_age'):
                self.last_age = "Unknown"
            if not hasattr(self, 'last_emotion'):
                self.last_emotion = "Unknown"
            if not hasattr(self, 'last_face_locations'):
                self.last_face_locations = []

            # Process every 4th frame
            if self.frame_counter % 4 == 0:
                # Resize frame for faster processing
                small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

                # Detect age and emotion
                self.last_age, self.last_emotion = self.detect_age_emotion(small_frame)

                # Convert to RGB for face detection
                rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

                # Detect faces
                face_locations = face_recognition.face_locations(rgb_small_frame)

                # Scale back face locations
                self.last_face_locations = [(top * 4, right * 4, bottom * 4, left * 4)
                                            for (top, right, bottom, left) in face_locations]

            # Draw face rectangles and labels
            for (top, right, bottom, left) in self.last_face_locations:
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

            cv2.putText(frame, f"Age: {self.last_age}", (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 1)
            cv2.putText(frame, f"Emotion: {self.last_emotion}", (10, 60),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 1)

            # Update capture
            self.most_recent_capture_arr = frame
            self.frame_counter += 1

            # Convert for display
            img_ = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            self.most_recent_capture_pil = Image.fromarray(img_)
            imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
            self._label.imgtk = imgtk
            self._label.configure(image=imgtk)

        # Repeat every 20 milliseconds
        self._label.after(20, self.process_webcam)

    def login(self):
        threading.Thread(target=self._login_thread, daemon=True).start()

    def _login_thread(self):
        temp_img_path = Path("./temp/temp_capture.jpg")
        temp_img_path.parent.mkdir(exist_ok=True, parents=True)

        try:
            # Verify camera frame
            if self.most_recent_capture_arr is None:
                raise ValueError("No frame captured from camera")

            # Save temporary image
            rgb_image = cv2.cvtColor(self.most_recent_capture_arr, cv2.COLOR_BGR2RGB)
            if not cv2.imwrite(str(temp_img_path), rgb_image):
                raise IOError(f"Failed to save image to {temp_img_path}")

            # Face recognition
            cmd = [
                'face_recognition',
                '--tolerance', '0.6',
                str(self.db_dir),
                str(temp_img_path)
            ]
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=15
            )

            if result.returncode != 0:
                raise ValueError(f"Face recognition failed: {result.stderr.strip()}")

            output = result.stdout.strip()
            if not output:
                raise ValueError("No face recognition output")

            parts = output.split(',')
            if len(parts) < 2:
                raise ValueError(f"Unexpected output format: {output}")

            name = parts[1].strip()
            if name in ['unknown_person', 'no_persons_found']:
                raise ValueError(f"Recognition failed: {name}")

            # Emotion detection
            try:
                _, emotion = self.detect_age_emotion(self.most_recent_capture_arr)
                emotion = emotion.lower()
            except Exception as e:
                logging.warning(f"Emotion detection failed: {str(e)}")
                emotion = "neutral"

            # Handle successful login
            welcome_msg = {
                'happy': f"Welcome {name}! You look happy today!",
                'sad': f"Welcome {name}. We hope your day gets better.",
                'angry': f"Welcome {name}. We appreciate your patience.",
            }.get(emotion, f"Welcome {name}!")

            self.main_window.after(0, lambda: util.msg_box(
                'Welcome',
                welcome_msg,
                icon_path="icons/shield.png"
            ))
            self.speak(welcome_msg)

            # Log attendance
            try:
                self._log_attendance(name, emotion)
            except Exception as e:
                self._log_to_text_backup(name, emotion, datetime.now().strftime("%H:%M:%S"))
                raise

        except Exception as e:
            error_msg = str(e) if str(e) else "Unknown error occurred"
            logging.error(f"Login error: {error_msg}")
            self.main_window.after(0, lambda: util.msg_box(
                'Login Failed',
                'Could not complete login process',
                icon_path="icons/x-button.png"
            ))
            self.speak("Login failed")

        finally:
            try:
                if temp_img_path.exists():
                    temp_img_path.unlink()
            except Exception as e:
                logging.warning(f"Cleanup error: {str(e)}")

    def register_new_user(self):
        self.register_new_user_window = tk.Toplevel(self.main_window)
        self.register_new_user_window.geometry("1200x520+370+120")
        self.register_new_user_window.title("FaceRec Register")

        self.accept_button_register_new_user_window = util.get_button(
            self.register_new_user_window,
            'Accept',
            'green',
            self.accept_register_new_user
        )
        self.accept_button_register_new_user_window.place(x=750, y=300)

        self.try_again_button_register_new_user_window = util.get_button(
            self.register_new_user_window,
            'Try again',
            'red',
            self.try_again_register_new_user
        )
        self.try_again_button_register_new_user_window.place(x=750, y=400)

        self.capture_label = util.get_img_label(self.register_new_user_window)
        self.capture_label.place(x=10, y=10, width=700, height=500)

        self.add_img_to_label(self.capture_label)

        self.entry_text_register_new_user = util.get_entry_text(self.register_new_user_window)
        self.entry_text_register_new_user.place(x=750, y=150)

        self.text_label_register_new_user = util.get_text_label(
            self.register_new_user_window,
            "Please, \ninput username: "
        )
        self.text_label_register_new_user.place(x=750, y=70)

        # Adding Hover Effects
        self.accept_button_register_new_user_window.bind("<Enter>", lambda e: self.on_enter(e, "darkgreen"))
        self.accept_button_register_new_user_window.bind("<Leave>", lambda e: self.on_leave(e, "green"))

        self.try_again_button_register_new_user_window.bind("<Enter>", lambda e: self.on_enter(e, "darkred"))
        self.try_again_button_register_new_user_window.bind("<Leave>", lambda e: self.on_leave(e, "red"))

    def try_again_register_new_user(self):
        self.register_new_user_window.destroy()

    def add_img_to_label(self, label):
        imgtk = ImageTk.PhotoImage(image=self.most_recent_capture_pil)
        label.imgtk = imgtk
        label.configure(image=imgtk)
        self.register_new_user_capture = self.most_recent_capture_arr.copy()

    def accept_register_new_user(self):
        # Disable the button to prevent multiple clicks
        self.accept_button_register_new_user_window.config(state=tk.DISABLED)

        name = self.entry_text_register_new_user.get(1.0, "end-1c")

        # Save the captured image
        cv2.imwrite(os.path.join(self.db_dir, '{}.jpg'.format(name)),
                    self.register_new_user_capture)

        # Log the registration
        self._log_attendance(name, "registration")

        # Close the registration window
        if hasattr(self, 'register_new_user_window') and self.register_new_user_window.winfo_exists():
            self.register_new_user_window.destroy()

        # Show success message
        self.main_window.after(0, lambda: util.msg_box('Success', 'User was registered successfully!'))

    def start(self):
        self.main_window.mainloop()

    def cleanup(self):
        """Proper cleanup on application exit"""
        # TTS cleanup
        if hasattr(self, 'engine'):
            with self.engine_lock:
                try:
                    self.engine.endLoop()
                except:
                    pass

        # Camera cleanup
        if hasattr(self, 'cap'):
            self.cap.release()

        # Excel cleanup
        if hasattr(self, 'workbook'):
            try:
                self.workbook.close()
            except:
                pass

        logging.info("Application cleaned up successfully")


if __name__ == "__main__":
    app = App()
    try:
        app.start()
    finally:
        app.cleanup()